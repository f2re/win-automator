from __future__ import annotations

import json
import os
import sqlite3
import tempfile
from pathlib import Path
from typing import Callable, List, Tuple


def _check(name: str, callback: Callable[[], None], results: List[Tuple[str, bool, str]]) -> None:
    try:
        callback()
        results.append((name, True, ""))
    except Exception as exc:
        results.append((name, False, "{}: {}".format(type(exc).__name__, exc)))


def run_smoke_test(gui: bool = False) -> dict:
    """Validate the packaged runtime without network or external test data."""
    results: List[Tuple[str, bool, str]] = []

    def import_runtime() -> None:
        import comtypes  # noqa: F401
        import openpyxl  # noqa: F401
        import pynput  # noqa: F401
        import pywinauto  # noqa: F401
        import tkinter  # noqa: F401
        import win32api  # noqa: F401
        from .debug_capture import DebugSession, DebugSink  # noqa: F401
        from .debug_ui import DebugCaptureApp  # noqa: F401

    _check("runtime-imports", import_runtime, results)

    def tcl_test() -> None:
        import tkinter
        interp = tkinter.Tcl()
        if not interp.eval("info patchlevel"):
            raise RuntimeError("Tcl runtime unavailable")

    _check("tcl", tcl_test, results)

    def win32_test() -> None:
        import win32api
        win32api.GetVersionEx()

    _check("win32", win32_test, results)

    def uia_test() -> None:
        from pywinauto import Desktop
        from pywinauto.controls.uiawrapper import UIAWrapper  # noqa: F401
        Desktop(backend="uia")

    _check("uia-backend", uia_test, results)

    with tempfile.TemporaryDirectory(prefix="win-automator-smoke-") as tmp:
        root = Path(tmp)

        def xlsx_test() -> None:
            from openpyxl import Workbook, load_workbook
            path = root / "smoke.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ФИО", "Отдел"])
            ws.append(["Иванов И.И.", "ОМН"])
            wb.save(str(path))
            wb.close()
            loaded = load_workbook(str(path), read_only=True, data_only=True)
            try:
                if loaded.active["A2"].value != "Иванов И.И.":
                    raise RuntimeError("XLSX roundtrip failed")
            finally:
                loaded.close()

        _check("xlsx", xlsx_test, results)

        def scenario_test() -> None:
            from .models import Scenario, Step, ValueSpec
            path = root / "scenario.json"
            scenario = Scenario(name="smoke", steps=[Step(action="key", key="{ENTER}", value=ValueSpec(literal=""))])
            scenario.save(path)
            loaded = Scenario.load(path)
            if loaded.name != "smoke" or len(loaded.steps) != 1:
                raise RuntimeError("Scenario serialization failed")

        _check("scenario", scenario_test, results)

        def checkpoint_test() -> None:
            from .storage import CheckpointDB
            db = CheckpointDB(root / "checkpoint.sqlite3")
            try:
                db.create_job("sample.xlsx", "Sheet1", "smoke", 2)
                row = db.latest_incomplete("sample.xlsx", "Sheet1", "smoke")
                if not row:
                    raise RuntimeError("Checkpoint roundtrip failed")
            finally:
                db.conn.close()

        _check("checkpoint", checkpoint_test, results)

        def debug_sink_test() -> None:
            from .debug_capture import ACTIVE_MARKER, DebugSink
            debug_root = root / "debug-smoke"
            debug_root.mkdir()
            (debug_root / ACTIVE_MARKER).write_text("smoke", encoding="utf-8")
            sink = DebugSink(debug_root, source="smoke")
            sink.log("smoke_event", ok=True)
            if not sink.events_path.exists() or "smoke_event" not in sink.events_path.read_text(encoding="utf-8"):
                raise RuntimeError("Debug JSONL writer failed")

        _check("debug-capture", debug_sink_test, results)

        def localappdata_test() -> None:
            from .storage import data_dir
            old_local = os.environ.get("LOCALAPPDATA")
            os.environ["LOCALAPPDATA"] = str(root / "appdata")
            try:
                target = data_dir() / "smoke.txt"
                target.write_text("ok", encoding="utf-8")
                if target.read_text(encoding="utf-8") != "ok":
                    raise RuntimeError("LOCALAPPDATA write roundtrip failed")
            finally:
                if old_local is None:
                    os.environ.pop("LOCALAPPDATA", None)
                else:
                    os.environ["LOCALAPPDATA"] = old_local

        _check("localappdata", localappdata_test, results)

        if gui:
            def gui_test() -> None:
                old_local = os.environ.get("LOCALAPPDATA")
                os.environ["LOCALAPPDATA"] = str(root / "gui-appdata")
                try:
                    from .ui import App
                    app = App()
                    app.withdraw()
                    app.update_idletasks()
                    app.update()
                    try:
                        app.db.conn.close()
                    except Exception:
                        pass
                    app.destroy()
                finally:
                    if old_local is None:
                        os.environ.pop("LOCALAPPDATA", None)
                    else:
                        os.environ["LOCALAPPDATA"] = old_local

            _check("gui", gui_test, results)

    failed = [item for item in results if not item[1]]
    return {
        "ok": not failed,
        "checks": [{"name": n, "ok": ok, "error": err} for n, ok, err in results],
    }


def write_report(path: Path, report: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
