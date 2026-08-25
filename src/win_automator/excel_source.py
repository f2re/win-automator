from __future__ import annotations

import datetime as dt
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

from openpyxl import load_workbook

from .debug_capture import DebugSink


def normalize(value: Any) -> Tuple[str, Any]:
    if value is None:
        return ("empty", "")
    if isinstance(value, bool):
        return ("bool", value)
    if isinstance(value, dt.datetime):
        return ("datetime", value.replace(microsecond=0))
    if isinstance(value, dt.date):
        return ("date", value)
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        try:
            return ("number", Decimal(str(value)).normalize())
        except InvalidOperation:
            pass

    text = re.sub(r"\s+", " ", str(value).strip())
    lowered = text.casefold()

    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%y"):
        try:
            return ("date", dt.datetime.strptime(text, fmt).date())
        except ValueError:
            continue

    numeric = text.replace(" ", "").replace(",", ".")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", numeric):
        try:
            return ("number", Decimal(numeric).normalize())
        except InvalidOperation:
            pass
    return ("text", lowered)


def infer_columns(value: Any, row: Dict[str, Any]) -> List[str]:
    needle = normalize(value)
    return [column for column, candidate in row.items() if normalize(candidate) == needle]


def unique_headers(values: Sequence[Any]) -> List[str]:
    result: List[str] = []
    counts: Dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = str(value).strip() if value not in (None, "") else "Столбец {}".format(index)
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else "{} ({})".format(base, counts[base]))
    return result


class ExcelSource:
    def __init__(self, path: Path) -> None:
        self.path = Path(path)
        self.debug = DebugSink.from_environment("excel")
        self._book = load_workbook(str(self.path), data_only=True, read_only=True)
        if self.debug:
            self.debug.log("excel_open", file_name=self.path.name, sheets=list(self._book.sheetnames))

    @property
    def sheets(self) -> List[str]:
        return list(self._book.sheetnames)

    def read(self, sheet: str, header_row: int = 1) -> Tuple[List[str], List[Dict[str, Any]]]:
        ws = self._book[sheet]
        iterator = ws.iter_rows(values_only=True)
        for _ in range(max(0, header_row - 1)):
            next(iterator, None)
        header_values = next(iterator, None)
        if header_values is None:
            if self.debug:
                self.debug.record_excel_schema(self.path, sheet, [], 0)
            return [], []
        headers = unique_headers(header_values)
        rows: List[Dict[str, Any]] = []
        for values in iterator:
            padded = list(values) + [None] * max(0, len(headers) - len(values))
            row = {headers[i]: padded[i] for i in range(len(headers))}
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        if self.debug:
            self.debug.record_excel_schema(self.path, sheet, headers, len(rows))
        return headers, rows

    def close(self) -> None:
        self._book.close()
        if self.debug:
            self.debug.log("excel_close", file_name=self.path.name)
