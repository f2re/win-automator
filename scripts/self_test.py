import importlib
import sqlite3
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))


def check(label, callback):
    try:
        callback()
        print("[OK] {}".format(label))
        return True
    except Exception as exc:
        print("[FAIL] {}: {}".format(label, exc))
        return False


def main():
    checks = []
    checks.append(check("Python {}".format(sys.version.split()[0]), lambda: None))
    checks.append(check("tkinter", lambda: importlib.import_module("tkinter")))
    checks.append(check("openpyxl", lambda: importlib.import_module("openpyxl")))
    checks.append(check("pywinauto", lambda: importlib.import_module("pywinauto")))
    checks.append(check("pynput", lambda: importlib.import_module("pynput")))
    checks.append(check("comtypes", lambda: importlib.import_module("comtypes")))
    checks.append(check("win32api", lambda: importlib.import_module("win32api")))

    def sqlite_test():
        conn = sqlite3.connect(":memory:")
        conn.execute("create table t(v integer)")
        conn.execute("insert into t values(1)")
        assert conn.execute("select v from t").fetchone()[0] == 1

    checks.append(check("SQLite", sqlite_test))

    def xlsx_test():
        from openpyxl import Workbook, load_workbook
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ФИО", "Отдел"])
            ws.append(["Иванов И.И.", "ОП"])
            wb.save(str(path))
            wb.close()
            wb2 = load_workbook(str(path), read_only=True, data_only=True)
            assert wb2.active["A2"].value == "Иванов И.И."
            wb2.close()

    checks.append(check("XLSX read/write", xlsx_test))
    print("\n{} / {} checks passed".format(sum(bool(x) for x in checks), len(checks)))
    return 0 if all(checks) else 1


if __name__ == "__main__":
    raise SystemExit(main())
