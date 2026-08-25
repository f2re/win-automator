import datetime as dt
import json
import subprocess
import sys
import time
from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from win_automator.excel_source import ExcelSource
from win_automator.executor import Executor
from win_automator.inspector import selector_from_wrapper
from win_automator.models import Scenario, Selector, Step, ValueSpec


pytestmark = pytest.mark.skipif(sys.platform != "win32", reason="Windows UI Automation integration test")


def _wait_for_file(path: Path, timeout: float = 15.0) -> None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        if path.exists() and path.stat().st_size:
            return
        time.sleep(0.1)
    raise AssertionError("Integration target did not write {}".format(path))


def _wait_for_window(desktop, process, timeout: float = 15.0):
    deadline = time.time() + timeout
    while time.time() < deadline:
        if process.poll() is not None:
            stdout, stderr = process.communicate(timeout=2)
            raise AssertionError(
                "Integration target exited before showing its window (code {}).\nstdout:\n{}\nstderr:\n{}".format(
                    process.returncode,
                    stdout.decode("utf-8", errors="replace"),
                    stderr.decode("utf-8", errors="replace"),
                )
            )
        window = desktop.window(title="WinAutomator Integration Target")
        if window.exists(timeout=0.2):
            window.wait("visible enabled", timeout=5)
            return window
        time.sleep(0.1)
    raise AssertionError("Integration target process is alive but its window is not visible")


def _captured(window, automation_id: str):
    wrapper = window.child_window(auto_id=automation_id).wrapper_object()
    return selector_from_wrapper(wrapper, "uia")


def test_excel_to_multi_window_application_end_to_end(tmp_path):
    # Workbook deliberately starts with a title row and contains typed dates and
    # numeric codes whose operator-visible format is learned from the first row.
    workbook_path = tmp_path / "integration.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"
    ws.merge_cells("A1:F1")
    ws["A1"] = "Служебная выгрузка"
    ws.append([
        "ФИО",
        "Дата рождения",
        "Отдел",
        "Код основной",
        "Код дополнительный",
        "Подтверждение",
    ])
    ws.append(["Иванов И.И.", dt.date(2020, 2, 1), "ОП", 7, 8, "TRAIN"])
    ws.append(["Петров П.П.", dt.date(2021, 4, 3), "ОМН", 12, 34, "OK"])
    wb.save(workbook_path)

    source = ExcelSource(workbook_path)
    try:
        header_row = source.detect_header_row("Сотрудники")
        headers, rows = source.read("Сотрудники", header_row=header_row)
    finally:
        source.close()
    assert header_row == 2
    assert headers[0] == "ФИО"
    assert [row["ФИО"] for row in rows] == ["Иванов И.И.", "Петров П.П."]

    output_path = tmp_path / "result.json"
    target_script = ROOT / "scripts" / "IntegrationTarget.ps1"
    process = subprocess.Popen(
        [
            "powershell.exe",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(target_script),
            "-OutputPath",
            str(output_path),
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )

    try:
        from pywinauto import Desktop

        desktop = Desktop(backend="uia")
        window = _wait_for_window(desktop, process)

        name_selector = _captured(window, "txtFullName")
        date_selector = _captured(window, "txtBirthDate")
        department_selector = _captured(window, "cmbDepartment")
        primary_selector = _captured(window, "txtPrimaryCode")
        secondary_selector = _captured(window, "txtSecondaryCode")
        next_selector = _captured(window, "btnNext")

        # Force the two code controls through the difficult fallback path: same
        # control type/class/name and no stable automation/control id. Relative
        # position must keep them distinct.
        for selector in (primary_selector, secondary_selector):
            selector.automation_id = ""
            selector.control_id = None
            selector.name = ""
            selector.parent_name = ""

        # The modal does not exist while the scenario is being assembled, so do
        # not invent an AutomationId for it. In WinForms, Control.Name is not a
        # guaranteed UIA AutomationId. The recorder-visible identity here is the
        # dialog title + process + accessible name + control type; the same
        # semantic fallback is what real recorded modal controls can use.
        dialog_confirm = Selector(
            backend="uia",
            window_title="WinAutomator Integration Dialog",
            process_name="powershell.exe",
            control_type="Edit",
            name="Подтверждение",
        )
        dialog_done = Selector(
            backend="uia",
            window_title="WinAutomator Integration Dialog",
            process_name="powershell.exe",
            control_type="Button",
            name="Готово",
        )

        scenario = Scenario(
            name="E2E",
            steps=[
                Step(
                    action="set_value",
                    target=name_selector,
                    value=ValueSpec(source="excel", column="ФИО", literal="Иванов И.И."),
                    description="ФИО",
                ),
                Step(
                    action="set_value",
                    target=date_selector,
                    value=ValueSpec(source="excel", column="Дата рождения", literal="01.02.2020"),
                    description="Дата",
                ),
                Step(
                    action="select",
                    target=department_selector,
                    value=ValueSpec(source="excel", column="Отдел", literal="ОП"),
                    description="Отдел",
                ),
                Step(
                    action="set_value",
                    target=primary_selector,
                    value=ValueSpec(source="excel", column="Код основной", literal="0007"),
                    description="Основной код",
                ),
                Step(
                    action="set_value",
                    target=secondary_selector,
                    value=ValueSpec(source="excel", column="Код дополнительный", literal="0008"),
                    description="Дополнительный код",
                ),
                Step(action="click", target=next_selector, description="Открыть модальное окно"),
                Step(
                    action="set_value",
                    target=dialog_confirm,
                    value=ValueSpec(source="excel", column="Подтверждение", literal="TRAIN"),
                    description="Подтверждение",
                ),
                Step(action="click", target=dialog_done, description="Завершить"),
            ],
        )

        result = Executor().run(scenario, rows[1])
        assert result.ok, str(result.error)
        assert result.completed_steps == 8
        _wait_for_file(output_path)

        payload = json.loads(output_path.read_text(encoding="utf-8-sig"))
        assert payload == {
            "full_name": "Петров П.П.",
            "birth_date": "03.04.2021",
            "department": "ОМН",
            "primary_code": "0012",
            "secondary_code": "0034",
            "confirm": "OK",
        }
    finally:
        if process.poll() is None:
            process.terminate()
            try:
                process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                process.kill()
