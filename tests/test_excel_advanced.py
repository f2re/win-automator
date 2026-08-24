import datetime as dt
import sys
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from win_automator.excel_source import ExcelSource, format_like_sample, infer_column


def test_detects_title_row_and_preserves_record_order(tmp_path):
    path = tmp_path / "input.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"
    ws.merge_cells("A1:D1")
    ws["A1"] = "Выгрузка сотрудников"
    ws.append(["ФИО", "Дата рождения", "Отдел", "Код"])
    ws.append(["Первый", dt.date(2020, 1, 2), "ОП", 7])
    ws.append([None, None, None, None])
    ws.append(["Второй", dt.date(2021, 3, 4), "ОМН", 12])
    wb.save(path)

    source = ExcelSource(path)
    try:
        header_row = source.detect_header_row("Данные")
        headers, rows = source.read("Данные", header_row=header_row)
    finally:
        source.close()

    assert header_row == 2
    assert headers == ["ФИО", "Дата рождения", "Отдел", "Код"]
    assert [row["ФИО"] for row in rows] == ["Первый", "Второй"]
    assert rows[1]["Код"] == 12


def test_duplicate_value_uses_semantic_target_hint_only_when_unique():
    row = {
        "Город рождения": "Москва",
        "Город регистрации": "Москва",
        "Страна": "Россия",
    }
    assert infer_column("Москва", row, hints=["Город рождения"]) == "Город рождения"
    assert infer_column("Москва", row, hints=["Город"]) is None


def test_format_like_training_preserves_dates_zeroes_and_decimal_comma():
    assert format_like_sample(dt.date(2026, 8, 24), "01.02.2020") == "24.08.2026"
    assert format_like_sample(12, "0007") == "0012"
    assert format_like_sample(Decimal("2.5"), "1,50") == "2,50"
    assert format_like_sample(Decimal("2.5"), "1.50") == "2.50"
